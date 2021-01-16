import sys
import getopt
import csv
import os
import shutil
import chardet
import tempfile
import openpyxl


def get_header_check(header, delimiter, header_valid):
    list_valid = header_valid.lower().split(';')
    list_header = header.lower().split(delimiter)
    missing = []
    unnecessary = []
    error = False
    unordered = False
    for i in range(len(list_valid)):
        if list_valid[i] not in list_header:
            missing.append(list_valid[i])
            error = True
    for i in range(len(list_header)):
        if list_header[i] not in list_valid:
            unnecessary.append(list_header[i])
            error = True
    if len(missing) == 0 and len(unnecessary) == 0:
        for i in range(len(list_valid)):
            if list_valid[i] != list_header[i]:
                unordered = True
                error = True
                break
    if error:
        result = 'KO:'
        if unordered:
            result = result + ' Desordenado'
        if len(missing) > 0:
            result = result + ' Faltan:' + ','.join(missing)
        if len(unnecessary) > 0:
            result = result + ' Sobran:' + ','.join(unnecessary)
        return result
    else:
        return 'OK'


def get_details_file(file_path, header=None):
    if header is None:
        details = {'encoding': '', 'confidence': 0.0, 'delimiter': '', 'rows': 0, 'header': ''}
    else:
        details = {'encoding': '', 'confidence': 0.0, 'delimiter': '', 'rows': 0, 'header': '', 'check_header': ''}
    if os.path.isfile(file_path):
        file_extension = os.path.splitext(file_path)[1].lower()
        if file_extension in ['.csv', '.txt']:
            with open(file_path, 'rb') as f:
                content_bytes = f.read()
            detected = chardet.detect(content_bytes)
            details['encoding'] = detected['encoding']
            details['confidence'] = detected['confidence']
            with open(file_path, 'r') as f:
                dialect = csv.Sniffer().sniff(f.readline())
                details['delimiter'] = dialect.delimiter
                f.seek(0)
                details['header'] = f.readline().strip()
                details['rows'] = sum(1 for row in f)
            if header is not None:
                details['check_header'] = get_header_check(details['header'], details['delimiter'], header)
        elif file_extension in ['.xls', '.xlsx']:
            with openpyxl.load_workbook(file_path, read_only=True) as wb:
                details['encoding'] = wb.encoding
                details['confidence'] = 1.0
                sh_header = None
                for h in range(1, wb.active.max_column + 1):
                    if sh_header is None:
                        sh_header = wb.active.cell(1, h).value
                    else:
                        sh_header = sh_header + ';' + wb.active.cell(1, h).value
                details['header'] = sh_header
                details['rows'] = wb.worksheets[0].max_row
            if header is not None:
                details['check_header'] = get_header_check(details['header'], ';', header)
    return details


def get_data_file(file_path):
    data_list_dict = []
    if os.path.isfile(file_path):
        file_extension = os.path.splitext(file_path)[1].lower()
        if file_extension in ['.csv', '.txt']:
            details = get_details_file(file_path)
            with open(file_path, 'r', encoding=details['encoding']) as csv_reader:
                reader = csv.reader(csv_reader, delimiter=details['delimiter'])
                header = None
                for row in reader:
                    row_dict = {}
                    if header is None:
                        header = row
                    else:
                        for i in range(len(header)):
                            row_dict[header[i]] = row[i]
                        data_list_dict.append(row_dict)
        elif file_extension in ['.xls', '.xlsx']:
            wb_obj = openpyxl.load_workbook(file_path)
            sh_obj = wb_obj.active
            max_row = sh_obj.max_row
            max_col = sh_obj.max_column
            header = []
            for r in range(1, max_row + 1):
                row_dict = {}
                for c in range(1, max_col + 1):
                    cl_obj = sh_obj.cell(row=r, column=c)
                    if r == 1:
                        header.append(cl_obj.value)
                    else:
                        row_dict[header[c - 1]] = cl_obj.value
                data_list_dict.append(row_dict)
            wb_obj.close()
    return data_list_dict


def set_data_csv(file_path, data_list_dict, to_utf8=False):
    if to_utf8:
        if os.path.isfile(file_path):
            details = get_details_file(file_path)
            if details['confidence'] >= 0.8:
                with open(file_path, 'rb') as f:
                    content_bytes = f.read()
                content_text = content_bytes.decode(details['encoding'])
                with tempfile.NamedTemporaryFile('w',
                                                 encoding='utf-8',
                                                 newline='',
                                                 dir=os.path.dirname(file_path),
                                                 delete=False) as f:
                    f.write(content_text)
                os.replace(f.name, file_path)
    else:
        with open(file_path, 'w', encoding='utf-8', newline='') as csv_writer:
            writer = csv.writer(csv_writer, delimiter=';', quoting=csv.QUOTE_MINIMAL)
            header = data_list_dict[0].keys()
            writer.writerow(header)
            for row in data_list_dict:
                row_list = []
                for head in header:
                    row_list.append(row[head])
                writer.writerow(row_list)


def get_file_path(path, filename):
    result = None
    if os.path.isdir(path):
        list_dir = os.listdir(path)
        for item in list_dir:
            if os.path.isdir(os.path.join(path, item)):
                result = get_file_path(os.path.join(path, item), filename)
                if result is not None:
                    break
            elif os.path.isfile(os.path.join(path, item)):
                if item.strip().lower() == filename.strip().lower():
                    result = os.path.join(path, item)
                    break
    return result


def copy_file(file_path, path):
    if os.path.isfile(file_path):
        file_path_copy = os.path.join(path, os.path.basename(file_path))
        if not os.path.isfile(file_path_copy):
            if not os.path.isdir(path):
                os.makedirs(path)
            shutil.copyfile(file_path, file_path_copy)


def version():
    print('1.0.0')


def usage():
    print('''
    Gather the files <input> of a list <field> in the same path <copy>, being able to be already 
    in the destination path or having to look for them in another <search>. Copy all the files in
    the destination path, validating if necessary their headers <header> and encoding, to transform
    them into csv with a common utf-8 format.
    
    Usage:
      python main.py --help
      python main.py --version
      python main.py -i <infile> -f <field> -c <copy>
      python main.py -i <infile> -f <field> -c <copy> -s <search>
      python main.py -i <infile> -f <field> -c <copy> -H <header>
      python main.py -i <infile> -f <field> -c <copy> -s <search> -H <header>
    
    Mandatory arguments to long options are mandatory for short options too.
      -i, --input=STRING file path with the list of files
      -f, --field=STRING name of field on the input file with list of files
      -c, --copy=STRING path destination to copy files
      -s, --search=STRING optional path to search files that not exists on path destination
      -H, --header=STRING optional header to validate with fields separated by ;
      -h, --help        display this help and exit
      -v, --version     display version and exit
    
    If <input>, <field> or <copy> are omitted, it exit program. When <search> is informed if file 
    not exists on <copy> search it on directory <search> and subdirectories. And when <header> is
    informed check if headers of files are valid. And then write a file log as <input> adding 
    'result' to filename adding data of files as encoding, headers, rows, ...
    ''')


def parse():
    infile = None
    field = None
    copy = None
    search = None
    header = None
    try:
        options, arguments = getopt.getopt(
            sys.argv[1:],                                                               # Arguments
            'i:f:c:s:H:hv',                                                             # Short option definitions
            ['input=', 'field=', 'copy=', 'search=', 'header=', 'help', 'version'])     # Long option definitions
    except getopt.GetoptError:
        usage()
        sys.exit()
    for o, a in options:
        if o in ('-v', '--version'):
            version()
            sys.exit()
        if o in ('-h', '--help'):
            usage()
            sys.exit()
        if o in ('-i', '--input'):
            infile = a
        if o in ('-f', '--field'):
            field = a
        if o in ('-c', '--copy'):
            copy = a
        if o in ('-s', '--search'):
            search = a
        if o in ('-H', '--header'):
            header = a
    if infile is None or field is None or copy is None:
        usage()
        sys.exit()
    return infile, field, copy, search, header


def main():
    print(parse())

#    file_list_files = None
#    field_name_file = None
#    path_to_copy = None
#    path_to_search = None
#    header_valid = None
#    try:
#        opts, args = getopt.getopt(argv, 'i:c:p:s:v:')
#    except getopt.GetoptError:
#        print('main.py'
#              ' -i <file_list_files>'
#              ' -c <field_name_file>'
#              ' -p <path_to_copy>'
#              ' -s <path_to_search>'
#              ' -H <header_valid>')
#        sys.exit(2)
#    for opt, arg in opts:
#        if opt == '-h':
#            print('''
#                  main.py
#                  -i file with list of necessary files
#                  -c field of file with name of necessary files
#                  -p path to save necessary files
#                  -s path to search files
#                  -v (optional) header valid on necessary files
#                  ''')
#            sys.exit()
#        elif opt in ('-i', '--file_list_files'):
#            file_list_files = arg
#        elif opt in ('-c', '--field_name_file'):
#            field_name_file = arg
#        elif opt in ('-p', '--path_to_copy'):
#            path_to_copy = arg
#        elif opt in ('-s', '--path_to_search'):
#            path_to_search = arg
#        elif opt in ('-v', '--header_valid'):
#            header_valid = arg

#    print('Input file is', file_list_files)
#    print('Column file is', field_name_file)
#    print('Path to files is', path_to_copy)
#    print('Path to search is', path_to_search)
#    print('Header valid is', header_valid)
#    path_to_search = r'C:\Users\fpx1vg\Desktop\BBDD'
#    path_to_copy = r'C:\Users\fpx1vg\Desktop\BBDD 2020'
#    csv_list_files = r'C:\Users\fpx1vg\Desktop\CARGAS_2020.csv'
#    csv_list_result = r'C:\Users\fpx1vg\Desktop\CARGAS_2020_RESULTADO.csv'
#    field_name_file = 'CARGA'
#    header_valid = 'LEAD_ID;E_CUPS;E_FIGURA_JUR;E_TITULAR_NOMBRE;E_TITULAR_APELLIDO1;E_TITULAR_APELLIDO2;E_TELEFONO;
#    E_TELEFONO2;E_TELEFONO3;E_TELEFONO4;E_EMAIL1;E_EMAIL2;E_MULTI;E_TANDA;E_PRIORIDAD;E_TARIFA;E_TENSION;
#    E_POTENCIA_CONTRATADA_P1;E_POTENCIA_CONTRATADA_P2;E_POTENCIA_CONTRATADA_P3;E_POTENCIA_NORMALIZADA;
#    E_PS_PROVINCIA;E_PS_MUNICIPIO;E_PS_POBLACION;E_PS_CODIGO_POSTAL;E_PS_TIPO_VIA;E_PS_CALLE;E_PS_NUMERO;E_PS_EDIFICIO;
#    E_PS_ESCALERA;E_PS_PORTAL;E_PS_PISO;E_PS_PUERTA;E_PS_ACLARADOR;E_DOC_NOMBRE;E_DOC_APELLIDO1;E_DOC_APELLIDO2;
#    E_DOC_PROVINCIA;E_DOC_MUNICIPIO;E_DOC_POBLACION;E_DOC_CODIGO_POSTAL;E_DOC_TIPO_VIA;E_DOC_CALLE;E_DOC_NUMERO;
#    E_DOC_EDIFICIO;E_DOC_PORTAL;E_DOC_ESCALERA;E_DOC_PISO;E_DOC_PUERTA;E_DOC_ACLARADOR;E_CLI_NOMBRE;E_CLI_APELLIDO1;
#    E_CLI_APELLIDO2;E_CLI_PROVINCIA;E_CLI_MUNICIPIO;E_CLI_POBLACION;E_CLI_CODIGO_POSTAL;E_CLI_TIPO_VIA;E_CLI_CALLE;
#    E_CLI_NUMERO;E_CLI_EDIFICIO;E_CLI_PORTAL;E_CLI_ESCALERA;E_CLI_PISO;E_CLI_PUERTA;E_CLI_ACLARADOR;G_CUPS;
#    G_FIGURA_JUR;G_TITULAR_NOMBRE;G_TITULAR_APELLIDO1;G_TITULAR_APELLIDO2;G_TELEFONO;G_TELEFONO2;G_TELEFONO3;
#    G_TELEFONO4;G_EMAIL1;G_EMAIL2;G_MULTI;G_TANDA;G_PRIORIDAD;G_TARIFA;G_PS_PROVINCIA;G_PS_MUNICIPIO;G_PS_POBLACION;
#    G_PS_CODIGO_POSTAL;G_PS_TIPO_VIA;G_PS_CALLE;G_PS_NUMERO;G_PS_EDIFICIO;G_PS_ESCALERA;G_PS_PORTAL;G_PS_PISO;
#    G_PS_PUERTA;G_PS_ACLARADOR;G_DOC_NOMBRE;G_DOC_APELLIDO1;G_DOC_APELLIDO2;G_DOC_PROVINCIA;G_DOC_MUNICIPIO;
#    G_DOC_POBLACION;G_DOC_CODIGO_POSTAL;G_DOC_TIPO_VIA;G_DOC_CALLE;G_DOC_NUMERO;G_DOC_EDIFICIO;G_DOC_PORTAL;
#    G_DOC_ESCALERA;G_DOC_PISO;G_DOC_PUERTA;G_DOC_ACLARADOR;G_CLI_NOMBRE;G_CLI_APELLIDO1;G_CLI_APELLIDO2;
#    G_CLI_PROVINCIA;G_CLI_MUNICIPIO;G_CLI_POBLACION;G_CLI_CODIGO_POSTAL;G_CLI_TIPO_VIA;G_CLI_CALLE;G_CLI_NUMERO;
#    G_CLI_EDIFICIO;G_CLI_PORTAL;G_CLI_ESCALERA;G_CLI_PISO;G_CLI_PUERTA;G_CLI_ACLARADOR;E_FECHA_CAMBIO_COMERC;
#    E_CONSUMO_ULT_12M;G_FECHA_CAMBIO_COMERC;G_CONSUMO_ULT_12M;ORIGEN;FRANJA_HORARIA_RECOMENDADA;E_PRODUCTO_RECOMENDADO;
#    G_PRODUCTO_RECOMENDADO;ANTIG_CONTRATO_ACTUAL;E_ZONA_GASIFICADA;SONDEO;RACIMO;CONSENTIMIENTO;E_COMERCIALIZADORA;
#    E_COMERCIALIZADORA_CUR;G_COMERCIALIZADORA;G_COMERCIALIZADORA_CUR;AUX_01;AUX_02;AUX_03;AUX_04;AUX_05;AUX_06;AUX_07;
#    AUX_08;AUX_09;AUX_10;AUX_11;AUX_12;AUX_13;AUX_14;AUX_15;AUX_16;AUX_17;AUX_18;AUX_19;AUX_20;AUX_21;AUX_22;AUX_23;
#    AUX_24;AUX_25;AUX_26;AUX_27;AUX_28;AUX_29;AUX_30;AUX_31;AUX_32;AUX_33;AUX_34;AUX_35;AUX_36;AUX_37;AUX_38;AUX_39;
#    AUX_40;AUX_41;AUX_42;AUX_43;AUX_44;AUX_45;AUX_46;AUX_47;AUX_48;AUX_49;AUX_50'

#    csv_list_dict = read_csv(csv_list_files)
#    counter = 1
#    for row in csv_list_dict:
#        try:
#            if field_name_file in row.keys():
#                print(f'Procesando el fichero {counter}: {row[field_name_file]} ...')
#                counter = counter + 1
#                path_file_search = search_file(path_to_search, row[field_name_file])
#                path_file_dst = os.path.join(path_to_copy, row[field_name_file])
#                if os.path.isfile(path_file_dst):
#                    row.update({'log': 'ficheiro existente'})
#                    row.update(data_file_details(path_file_dst))
#                    row.update({'header': check_header(row['header'], header_valid)})
#                elif os.path.isfile(path_file_search):
#                    copy_file(path_file_search, path_to_copy)
#                    row.update({'log': 'ficheiro copiado'})
#                    row.update(data_file_details(path_file_dst))
#                    row.update({'header': check_header(row['header'], header_valid)})
#                else:
#                    row.update({'log': 'fichero non existe'})
#                    row.update(data_file_details(''))
#            else:
#                raise Exception(f'Non existe o campo {field_name_file} no listado de ficheiros')


#        except Exception as e:
#            row.update({'log': f'ficheiro provoca un erro {e.__str__()}'})
#            row.update(data_file_details(''))
#            pass

# write_csv(csv_list_result, csv_list_dict)


if __name__ == '__main__':
    main()
