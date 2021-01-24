import sys
import getopt
import csv
import os
import shutil
import chardet
import tempfile
import openpyxl
import logging


def get_header_check(header, delimiter, header_valid):
    list_valid = header_valid.lower().split(';')
    list_header = header.lower().split(delimiter)
    missing = []
    unnecessary = []
    error = False
    unordered = False
    for i in range(len(list_valid)):
        if list_valid[i] not in list_header:
            missing.append(list_valid[i])
            error = True
    for i in range(len(list_header)):
        if list_header[i] not in list_valid:
            unnecessary.append(list_header[i])
            error = True
    if len(missing) == 0 and len(unnecessary) == 0:
        for i in range(len(list_valid)):
            if list_valid[i] != list_header[i]:
                unordered = True
                error = True
                break
    if error:
        result = 'KO:'
        if unordered:
            result = result + ' Desordenado'
        if len(missing) > 0:
            result = result + ' Faltan:' + ','.join(missing)
        if len(unnecessary) > 0:
            result = result + ' Sobran:' + ','.join(unnecessary)
        return result
    else:
        return 'OK'


def get_details_file(file_path, header=None):
    if header is None:
        details = {'extension': '', 'encoding': '', 'confidence': 0.0, 'delimiter': '', 'rows': 0, 'header': ''}
    else:
        details = {'extension': '', 'encoding': '', 'confidence': 0.0, 'delimiter': '', 'rows': 0, 'header': '', 'check_header': ''}
    if os.path.isfile(file_path):
        details['extension'] = os.path.splitext(file_path)[1].lower()
        if details['extension'] in ['.csv', '.txt']:
            with open(file_path, 'rb') as f:
                content_bytes = f.read()
            detected = chardet.detect(content_bytes)
            details['encoding'] = detected['encoding']
            details['confidence'] = detected['confidence']
            with open(file_path, 'r') as f:
                dialect = csv.Sniffer().sniff(f.readline())
                details['delimiter'] = dialect.delimiter
                f.seek(0)
                details['header'] = f.readline().strip()
                details['rows'] = sum(1 for row in f)
            if header is not None:
                details['check_header'] = get_header_check(details['header'], details['delimiter'], header)
        elif details['extension'] in ['.xls', '.xlsx']:
            with openpyxl.load_workbook(file_path, read_only=True) as wb:
                details['encoding'] = wb.encoding
                details['confidence'] = 1.0
                sh_header = None
                for h in range(1, wb.active.max_column + 1):
                    if sh_header is None:
                        sh_header = wb.active.cell(1, h).value
                    else:
                        sh_header = sh_header + ';' + wb.active.cell(1, h).value
                details['header'] = sh_header
                details['rows'] = wb.worksheets[0].max_row
            if header is not None:
                details['check_header'] = get_header_check(details['header'], ';', header)
    return details


def get_data_file(file_path):
    data_list_dict = []
    if os.path.isfile(file_path):
        file_extension = os.path.splitext(file_path)[1].lower()
        if file_extension in ['.csv', '.txt']:
            details = get_details_file(file_path)
            with open(file_path, 'r', encoding=details['encoding']) as csv_reader:
                reader = csv.reader(csv_reader, delimiter=details['delimiter'])
                header = None
                for row in reader:
                    row_dict = {}
                    if header is None:
                        header = row
                    else:
                        for i in range(len(header)):
                            row_dict[header[i]] = row[i]
                        data_list_dict.append(row_dict)
        elif file_extension in ['.xls', '.xlsx']:
            wb_obj = openpyxl.load_workbook(file_path)
            sh_obj = wb_obj.active
            max_row = sh_obj.max_row
            max_col = sh_obj.max_column
            header = []
            for r in range(1, max_row + 1):
                row_dict = {}
                for c in range(1, max_col + 1):
                    cl_obj = sh_obj.cell(row=r, column=c)
                    if r == 1:
                        header.append(cl_obj.value)
                    else:
                        row_dict[header[c - 1]] = cl_obj.value
                data_list_dict.append(row_dict)
            wb_obj.close()
    return data_list_dict


def set_data_csv(file_path, data_list_dict={}, to_utf8=False):
    if to_utf8:
        if os.path.isfile(file_path):
            details = get_details_file(file_path)
            if details['confidence'] >= 0.8:
                with open(file_path, 'rb') as f:
                    content_bytes = f.read()
                content_text = content_bytes.decode(details['encoding'])
                with tempfile.NamedTemporaryFile('w',
                                                 encoding='utf-8',
                                                 newline='',
                                                 dir=os.path.dirname(file_path),
                                                 delete=False) as f:
                    f.write(content_text)
                os.replace(f.name, file_path)
    else:
        with open(file_path, 'w', encoding='utf-8', newline='') as csv_writer:
            writer = csv.writer(csv_writer, delimiter=';', quoting=csv.QUOTE_MINIMAL)
            header = data_list_dict[0].keys()
            writer.writerow(header)
            for row in data_list_dict:
                row_list = []
                for head in header:
                    row_list.append(row[head])
                writer.writerow(row_list)


def get_file_path(path, filename):
    result = None
    if os.path.isdir(path):
        list_dir = os.listdir(path)
        for item in list_dir:
            if os.path.isdir(os.path.join(path, item)):
                result = get_file_path(os.path.join(path, item), filename)
                if result is not None:
                    break
            elif os.path.isfile(os.path.join(path, item)):
                if item.strip().lower() == filename.strip().lower():
                    result = os.path.join(path, item)
                    break
    return result


def set_log():
    logger = logging.getLogger('et-csv')
    ch = logging.StreamHandler()
    ch.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
    logger.addHandler(ch)
    logger.setLevel(logging.DEBUG)

    return logger


def copy_file(file_path, path):
    if os.path.isfile(file_path):
        file_path_copy = os.path.join(path, os.path.basename(file_path))
        if not os.path.isfile(file_path_copy):
            if not os.path.isdir(path):
                os.makedirs(path)
            shutil.copyfile(file_path, file_path_copy)


def version():
    print('1.0.0')


def usage():
    print('''
    Gather the files <input> of a list <field> in the same path <copy>, being able to be already 
    in the destination path or having to look for them in another <search>. Copy all the files in
    the destination path, validating if necessary their headers <header> and encoding, to transform
    them into csv with a common utf-8 format.
    
    Usage:
      python et-csv.py --help
      python et-csv.py --version
      python et-csv.py -i <infile> -f <field> -c <copy>
      python et-csv.py -i <infile> -f <field> -c <copy> -s <search>
      python et-csv.py -i <infile> -f <field> -c <copy> -H <header>
      python et-csv.py -i <infile> -f <field> -c <copy> -s <search> -H <header>
    
    Mandatory arguments to long options are mandatory for short options too.
      -i, --input=STRING file path with the list of files
      -f, --field=STRING name of field on the input file with list of files
      -c, --copy=STRING path destination to copy files
      -s, --search=STRING optional path to search files that not exists on path destination
      -H, --header=STRING optional header to validate with fields separated by ;
      -h, --help        display this help and exit
      -v, --version     display version and exit
    
    If <input>, <field> or <copy> are omitted, it exit program. When <search> is informed if file 
    not exists on <copy> search it on directory <search> and subdirectories. And when <header> is
    informed check if headers of files are valid. And then write a file log as <input> adding 
    'result' to filename adding data of files as encoding, headers, rows, ...
    ''')


def parse():
    infile = None
    field = None
    copy = None
    search = None
    header = None
    try:
        options, arguments = getopt.getopt(
            sys.argv[1:],                                                               # Arguments
            'i:f:c:s:H:hv',                                                             # Short option definitions
            ['input=', 'field=', 'copy=', 'search=', 'header=', 'help', 'version'])     # Long option definitions
    except getopt.GetoptError:
        usage()
        sys.exit()
    for o, a in options:
        if o in ('-v', '--version'):
            version()
            sys.exit()
        if o in ('-h', '--help'):
            usage()
            sys.exit()
        if o in ('-i', '--input'):
            infile = a
        if o in ('-f', '--field'):
            field = a
        if o in ('-c', '--copy'):
            copy = a
        if o in ('-s', '--search'):
            search = a
        if o in ('-H', '--header'):
            header = a
    if infile is None or field is None or copy is None:
        usage()
        sys.exit()
    return {'infile': infile, 'field': field, 'copy': copy, 'search': search, 'header': header}


def main():
    log = set_log()
    params = parse()
    log.info('Start script')
    infile = params['infile']
    field = params['field']
    copy = params['copy']
    search = params['search']
    header = params['header']
    log.info(f'Read input file: {infile}')
    in_list_dict = get_data_file(infile)
    if len(in_list_dict) == 0:
        log.warning(f'No data on input file {infile}')
    counter = 1
    for row in in_list_dict:
        try:
            if field in row.keys():
                log.info(f'Searching file {counter}/{len(in_list_dict)} {row[field]}')
                file_copy = get_file_path(copy, row[field])
                if search is not None:
                    file_search = get_file_path(search, row[field])
                if os.path.isfile(file_copy):
                    log.info(f'File already exists')
                    row.update({'log': 'file exists'})
                elif os.path.isfile(file_search):
                    copy_file(file_search, copy)
                    log.info(f'File copied')
                    row.update({'log': 'file copied'})
                else:
                    log.info(f'File not found')
                    row.update({'log': 'file not found'})
                row.update(get_details_file(file_copy, header))
            else:
                log.warning(f'Not found field {field} on in file {infile}')
                break
            if row['extension'] in ['.csv', '.txt'] and row['encoding'] not in ['utf-8', 'ascii']:
                log.info(f'Convert encode file from {row["encoding"]} to utf-8')
                set_data_csv(file_copy, to_utf8=True)
                row.update(get_details_file(file_copy, header))
                log.info(f'Converted to {row["encoding"]}')
            counter = counter + 1
        except Exception as e:
            log.error(e.__str__())
            pass
    set_data_csv(f'{os.path.splitext(infile)[0]}_result.csv', in_list_dict)


if __name__ == '__main__':
    main()
